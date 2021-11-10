import openpyxl

excel = openpyxl.load_workbook("Location")
sheet = excel.active
#reading data from sheet
cell = sheet.cell(row=3, column=2)
print(cell.value)
print(sheet.max_row)
print(sheet.max_column)
#writing data into sheet
sheet.cell(row=3, column=4).value = "gfhgh"
list1 = []
for i in range(1, sheet.max_row+1):
    dict1 = {}
    for j in range(1, sheet.max_column):
        dict1[sheet.cell(row=1, column=j)] = sheet.cell(row=i, column=j)
    list1.append(dict1)
print(list1)